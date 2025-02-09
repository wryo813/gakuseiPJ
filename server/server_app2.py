import socket
import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

HOST = '0.0.0.0'
PORT = 50007
EXCEL_FILE = "log.xlsx"

def append_to_excel(
    time_str, 
    x_val, y_val, 
    servo1_val, servo2_val, servo3_val, servo4_val, 
    ip_addr, port
):
    file_exists = os.path.exists(EXCEL_FILE)

    if file_exists:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Time", 
            "x", 
            "y", 
            "servo1_angle", 
            "servo2_angle", 
            "servo3_angle", 
            "servo4_angle", 
            "IP", 
            "Port"
        ])

    ws.append([
        time_str,
        x_val,
        y_val,
        servo1_val,
        servo2_val,
        servo3_val,
        servo4_val,
        ip_addr,
        port
    ])

    wb.save(EXCEL_FILE)

def main():
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind((HOST, PORT))
        s.listen(1)
        print(f"Server listening on port {PORT}...")

        while True:
            conn, addr = s.accept()
            with conn:
                print("Connected by", addr)
                data = conn.recv(1024)
                if not data:
                    break

                message = json.loads(data.decode('utf-8'))
                print("Received:", message)

                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # ここで int に変換
                x_val      = int(message.get("x", 0))
                y_val      = int(message.get("y", 0))
                servo1_val = int(message.get("servo1_angle", 0))
                servo2_val = int(message.get("servo2_angle", 0))
                servo3_val = int(message.get("servo3_angle", 0))
                servo4_val = int(message.get("servo4_angle", 0))

                append_to_excel(
                    time_str=now_str,
                    x_val=x_val,
                    y_val=y_val,
                    servo1_val=servo1_val,
                    servo2_val=servo2_val,
                    servo3_val=servo3_val,
                    servo4_val=servo4_val,
                    ip_addr=addr[0],
                    port=addr[1]
                )

                response = {
                    "x": x_val,
                    "y": y_val,
                    "servo1_angle": servo1_val,
                    "servo2_angle": servo2_val,
                    "servo3_angle": servo3_val,
                    "servo4_angle": servo4_val
                }
                conn.sendall(json.dumps(response).encode('utf-8'))
                print("Sent back:", response)

if __name__ == "__main__":
    main()
